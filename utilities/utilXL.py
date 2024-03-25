import openpyxl
from openpyxl.workbook import Workbook


def getRows(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColumns(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


# read single value in a cell of excell
def readData(file, sheetName, rowno, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowno, column=columnno).value


# write single value in a cell of excell
def writeData(file, sheetName, rowno, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    workbook.save(file)
    return sheet.cell(row=rowno, column=columnno, value=data)


# write list data into excel in rows
def data_to_excel(data, filename):
    # Create a new Excel workbook
    wb = Workbook()
    # Get the active worksheet
    ws = wb.active
    # Write the data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook to a file
    wb.save(filename)


def read_all_dataInList(file_name, column_a, column_b):
    usernames = []
    passwords = []
    excel_path = f"/home/mubashar4603/PycharmProjects/lahebo-prodTest/TestData/{file_name}"

    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(excel_path)
        # Get the active worksheet
        worksheet = workbook.active

        # Extract data from the specified column, starting from the second row
        for cell_a, cell_b in zip(worksheet[column_a][1:], worksheet[column_b][1:]):
            usernames.append(cell_a.value)
            passwords.append(cell_b.value)

    except FileNotFoundError:
        print(f"File '{excel_path}' not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

    return usernames, passwords


# Example usage:
# username, password = read_all_dataInList('credential.xlsx', 'A', 'B')
# print("Data from Excel:", type(username))
# print(username, password)
